# app.py
import os
import io
import uuid
import json
import calendar
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, send_file, Response
from flask_cors import CORS
from supabase import create_client, Client
from weasyprint import HTML, CSS
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

app = Flask(__name__, static_folder='static', template_folder='templates')
CORS(app)

# Supabase configuratie
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Routes voor de webinterface
@app.route('/')
def index():
    return render_template('index.html')

# API Routes voor clienten
@app.route('/api/clienten', methods=['GET'])
def get_clienten():
    try:
        response = supabase.table('clienten').select('*').execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clienten', methods=['POST'])
def add_client():
    try:
        client_data = request.json
        response = supabase.table('clienten').insert(client_data).execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clienten/<id>', methods=['PUT'])
def update_client(id):
    try:
        client_data = request.json
        response = supabase.table('clienten').update(client_data).eq('id', id).execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/clienten/<id>', methods=['DELETE'])
def delete_client(id):
    try:
        response = supabase.table('clienten').delete().eq('id', id).execute()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API Routes voor behandelingen
@app.route('/api/behandelingen_types', methods=['GET'])
def get_behandelingen_types():
    try:
        response = supabase.table('behandelingen_types').select('*').execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/behandelingen_types', methods=['POST'])
def add_behandeling_type():
    try:
        behandeling_data = request.json
        response = supabase.table('behandelingen_types').insert(behandeling_data).execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/behandelingen_types/<id>', methods=['PUT'])
def update_behandeling_type(id):
    try:
        behandeling_data = request.json
        response = supabase.table('behandelingen_types').update(behandeling_data).eq('id', id).execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/behandelingen_types/<id>', methods=['DELETE'])
def delete_behandeling_type(id):
    try:
        response = supabase.table('behandelingen_types').delete().eq('id', id).execute()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API Routes voor uitgevoerde behandelingen
@app.route('/api/behandelingen', methods=['GET'])
def get_behandelingen():
    try:
        response = supabase.table('behandelingen').select('*,clienten(*),behandelingen_types(*)').execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/behandelingen', methods=['POST'])
def add_behandeling():
    try:
        behandeling_data = request.json
        response = supabase.table('behandelingen').insert(behandeling_data).execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/behandelingen/<id>', methods=['PUT'])
def update_behandeling(id):
    try:
        behandeling_data = request.json
        response = supabase.table('behandelingen').update(behandeling_data).eq('id', id).execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/behandelingen/<id>', methods=['DELETE'])
def delete_behandeling(id):
    try:
        response = supabase.table('behandelingen').delete().eq('id', id).execute()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API Routes voor facturatie
@app.route('/api/facturen/generate', methods=['POST'])
def generate_facturen():
    try:
        factuur_data = request.json
        start_datum = factuur_data.get('start_datum')
        eind_datum = factuur_data.get('eind_datum')
        
        # Haal behandelingen op binnen de opgegeven periode
        behandelingen = supabase.table('behandelingen')\
            .select('*,clienten(*),behandelingen_types(*)')\
            .gte('datum', start_datum)\
            .lte('datum', eind_datum)\
            .eq('gefactureerd', False)\
            .execute()
        
        # Groepeer behandelingen per client
        behandelingen_per_client = {}
        for behandeling in behandelingen.data:
            client_id = behandeling['client_id']
            if client_id not in behandelingen_per_client:
                behandelingen_per_client[client_id] = []
            behandelingen_per_client[client_id].append(behandeling)
        
        # Genereer facturen voor elke client
        gegenereerde_facturen = []
        for client_id, client_behandelingen in behandelingen_per_client.items():
            if not client_behandelingen:
                continue
                
            # Bereken totaalbedrag
            totaal_bedrag = 0
            for behandeling in client_behandelingen:
                behandeling_type = behandeling['behandelingen_types']
                if behandeling_type['prijs_type'] == 'per_uur':
                    totaal_bedrag += behandeling_type['prijs'] * behandeling['aantal_uren']
                else:  # per_sessie of anders
                    totaal_bedrag += behandeling_type['prijs']
            
            # Maak factuur aan
            factuur_nummer = f"F{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6]}"
            factuur = {
                'factuur_nummer': factuur_nummer,
                'client_id': client_id,
                'datum': datetime.now().strftime('%Y-%m-%d'),
                'start_datum': start_datum,
                'eind_datum': eind_datum,
                'totaal_bedrag': totaal_bedrag,
                'betaald': False
            }
            
            # Sla factuur op in database
            factuur_response = supabase.table('facturen').insert(factuur).execute()
            factuur_id = factuur_response.data[0]['id']
            
            # Update behandelingen als gefactureerd
            for behandeling in client_behandelingen:
                supabase.table('behandelingen')\
                    .update({'gefactureerd': True, 'factuur_id': factuur_id})\
                    .eq('id', behandeling['id'])\
                    .execute()
            
            # Voeg factuurdetails toe voor de response
            factuur['behandelingen'] = client_behandelingen
            factuur['client'] = client_behandelingen[0]['clienten']
            gegenereerde_facturen.append(factuur)
        
        return jsonify({
            'success': True, 
            'facturen': gegenereerde_facturen, 
            'aantal_facturen': len(gegenereerde_facturen)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/facturen', methods=['GET'])
def get_facturen():
    try:
        response = supabase.table('facturen').select('*,clienten(*)').execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/facturen/<id>', methods=['GET'])
def get_factuur(id):
    try:
        factuur = supabase.table('facturen').select('*,clienten(*)').eq('id', id).execute()
        behandelingen = supabase.table('behandelingen')\
            .select('*,behandelingen_types(*)')\
            .eq('factuur_id', id)\
            .execute()
        
        result = factuur.data[0] if factuur.data else {}
        result['behandelingen'] = behandelingen.data if behandelingen.data else []
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/facturen/<id>/betaald', methods=['PUT'])
def mark_factuur_betaald(id):
    try:
        response = supabase.table('facturen').update({'betaald': True}).eq('id', id).execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# PDF Factuurgeneratie met WeasyPrint
@app.route('/api/facturen/<id>/pdf', methods=['GET'])
def generate_factuur_pdf(id):
    try:
        # Haal factuurgegevens op
        factuur = supabase.table('facturen').select('*,clienten(*)').eq('id', id).execute()
        behandelingen = supabase.table('behandelingen')\
            .select('*,behandelingen_types(*)')\
            .eq('factuur_id', id)\
            .execute()
        
        if not factuur.data:
            return jsonify({'error': 'Factuur niet gevonden'}), 404
        
        factuur_data = factuur.data[0]
        factuur_data['behandelingen'] = behandelingen.data if behandelingen.data else []
        
        # Haal factuursjabloon op
        sjabloon_response = supabase.table('factuur_sjablonen').select('*').eq('actief', True).execute()
        if not sjabloon_response.data:
            # Gebruik standaardsjabloon als er geen actief is
            sjabloon_html = """
            <html>
                <head>
                    <style>
                        body { font-family: Arial, sans-serif; }
                        .factuur-header { text-align: center; margin-bottom: 20px; }
                        .factuur-details { margin-bottom: 20px; }
                        .behandelingen-tabel { width: 100%; border-collapse: collapse; }
                        .behandelingen-tabel th, .behandelingen-tabel td { 
                            border: 1px solid #ddd; padding: 8px; text-align: left; 
                        }
                        .totaal { text-align: right; font-weight: bold; margin-top: 20px; }
                    </style>
                </head>
                <body>
                    <div class="factuur-header">
                        <h1>Factuur</h1>
                        <p>Factuurnummer: {{ factuur.factuur_nummer }}</p>
                        <p>Datum: {{ factuur.datum }}</p>
                    </div>
                    
                    <div class="factuur-details">
                        <h3>Therapeutgegevens:</h3>
                        <p>Naam Praktijk</p>
                        <p>Adres Praktijk</p>
                        <p>Telefoonnummer</p>
                        <p>E-mail</p>
                        
                        <h3>Klantgegevens:</h3>
                        <p>{{ factuur.clienten.naam }}</p>
                        <p>{{ factuur.clienten.adres }}</p>
                        <p>{{ factuur.clienten.postcode }} {{ factuur.clienten.plaats }}</p>
                    </div>
                    
                    <h3>Behandelingen</h3>
                    <table class="behandelingen-tabel">
                        <thead>
                            <tr>
                                <th>Datum</th>
                                <th>Behandeling</th>
                                <th>Aantal</th>
                                <th>Prijs</th>
                                <th>Bedrag</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for behandeling in factuur.behandelingen %}
                            <tr>
                                <td>{{ behandeling.datum }}</td>
                                <td>{{ behandeling.behandelingen_types.naam }}</td>
                                <td>
                                    {% if behandeling.behandelingen_types.prijs_type == 'per_uur' %}
                                        {{ behandeling.aantal_uren }} uur
                                    {% else %}
                                        1
                                    {% endif %}
                                </td>
                                <td>€ {{ '{:.2f}'.format(behandeling.behandelingen_types.prijs) }}</td>
                                <td>
                                    {% if behandeling.behandelingen_types.prijs_type == 'per_uur' %}
                                        € {{ '{:.2f}'.format(behandeling.behandelingen_types.prijs * behandeling.aantal_uren) }}
                                    {% else %}
                                        € {{ '{:.2f}'.format(behandeling.behandelingen_types.prijs) }}
                                    {% endif %}
                                </td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                    
                    <div class="totaal">
                        <p>Totaalbedrag: € {{ '{:.2f}'.format(factuur.totaal_bedrag) }}</p>
                    </div>
                    
                    <div class="betaalinstructies">
                        <h3>Betaalinstructies:</h3>
                        <p>Graag het totaalbedrag binnen 14 dagen overmaken naar:</p>
                        <p>IBAN: NL00 BANK 0123 4567 89</p>
                        <p>t.n.v. Naam Praktijk</p>
                        <p>o.v.v. factuurnummer {{ factuur.factuur_nummer }}</p>
                    </div>
                </body>
            </html>
            """
        else:
            sjabloon_html = sjabloon_response.data[0]['html_content']
        
        # Render factuur HTML met Jinja2
        from jinja2 import Template
        template = Template(sjabloon_html)
        rendered_html = template.render(factuur=factuur_data)
        
        # Genereer PDF
        pdf = HTML(string=rendered_html).write_pdf()
        
        # Return PDF als download
        return Response(
            pdf,
            mimetype="application/pdf",
            headers={
                "Content-Disposition": f"attachment;filename=Factuur-{factuur_data['factuur_nummer']}.pdf"
            }
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Excel export
@app.route('/api/export/excel', methods=['POST'])
def export_to_excel():
    try:
        export_data = request.json
        start_datum = export_data.get('start_datum')
        eind_datum = export_data.get('eind_datum')
        
        # Haal alle benodigde data op
        clienten = supabase.table('clienten').select('*').execute()
        behandelingen_types = supabase.table('behandelingen_types').select('*').execute()
        behandelingen = supabase.table('behandelingen')\
            .select('*,clienten(*),behandelingen_types(*)')\
            .gte('datum', start_datum)\
            .lte('datum', eind_datum)\
            .execute()
        facturen = supabase.table('facturen')\
            .select('*,clienten(*)')\
            .gte('datum', start_datum)\
            .lte('datum', eind_datum)\
            .execute()
            
        # Maak workbook
        wb = Workbook()
        
        # Werkblad voor clienten
        ws_clienten = wb.active
        ws_clienten.title = "Clienten"
        
        # Voeg headers toe
        headers = ['ID', 'Naam', 'Adres', 'Postcode', 'Plaats', 'Telefoon', 'Email']
        for col_num, header in enumerate(headers, 1):
            cell = ws_clienten.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Voeg data toe
        for row_num, client in enumerate(clienten.data, 2):
            ws_clienten.cell(row=row_num, column=1).value = client.get('id')
            ws_clienten.cell(row=row_num, column=2).value = client.get('naam')
            ws_clienten.cell(row=row_num, column=3).value = client.get('adres')
            ws_clienten.cell(row=row_num, column=4).value = client.get('postcode')
            ws_clienten.cell(row=row_num, column=5).value = client.get('plaats')
            ws_clienten.cell(row=row_num, column=6).value = client.get('telefoon')
            ws_clienten.cell(row=row_num, column=7).value = client.get('email')
        
        # Werkblad voor behandelingen types
        ws_behandeling_types = wb.create_sheet("Behandeling Types")
        
        # Headers
        headers = ['ID', 'Naam', 'Beschrijving', 'Prijs', 'Prijs Type']
        for col_num, header in enumerate(headers, 1):
            cell = ws_behandeling_types.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Data
        for row_num, behandeling_type in enumerate(behandelingen_types.data, 2):
            ws_behandeling_types.cell(row=row_num, column=1).value = behandeling_type.get('id')
            ws_behandeling_types.cell(row=row_num, column=2).value = behandeling_type.get('naam')
            ws_behandeling_types.cell(row=row_num, column=3).value = behandeling_type.get('beschrijving')
            ws_behandeling_types.cell(row=row_num, column=4).value = behandeling_type.get('prijs')
            ws_behandeling_types.cell(row=row_num, column=5).value = behandeling_type.get('prijs_type')
        
        # Werkblad voor behandelingen
        ws_behandelingen = wb.create_sheet("Behandelingen")
        
        # Headers
        headers = ['ID', 'Datum', 'Client', 'Behandeling', 'Aantal Uren', 'Prijs Type', 'Prijs', 'Bedrag', 'Gefactureerd', 'Factuur ID']
        for col_num, header in enumerate(headers, 1):
            cell = ws_behandelingen.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Data
        for row_num, behandeling in enumerate(behandelingen.data, 2):
            ws_behandelingen.cell(row=row_num, column=1).value = behandeling.get('id')
            ws_behandelingen.cell(row=row_num, column=2).value = behandeling.get('datum')
            ws_behandelingen.cell(row=row_num, column=3).value = behandeling.get('clienten', {}).get('naam')
            ws_behandelingen.cell(row=row_num, column=4).value = behandeling.get('behandelingen_types', {}).get('naam')
            ws_behandelingen.cell(row=row_num, column=5).value = behandeling.get('aantal_uren')
            
            prijs_type = behandeling.get('behandelingen_types', {}).get('prijs_type')
            prijs = behandeling.get('behandelingen_types', {}).get('prijs', 0)
            
            ws_behandelingen.cell(row=row_num, column=6).value = prijs_type
            ws_behandelingen.cell(row=row_num, column=7).value = prijs
            
            if prijs_type == 'per_uur':
                bedrag = prijs * behandeling.get('aantal_uren', 0)
            else:
                bedrag = prijs
                
            ws_behandelingen.cell(row=row_num, column=8).value = bedrag
            ws_behandelingen.cell(row=row_num, column=9).value = 'Ja' if behandeling.get('gefactureerd') else 'Nee'
            ws_behandelingen.cell(row=row_num, column=10).value = behandeling.get('factuur_id')
        
        # Werkblad voor facturen
        ws_facturen = wb.create_sheet("Facturen")
        
        # Headers
        headers = ['ID', 'Factuurnummer', 'Datum', 'Client', 'Periode Start', 'Periode Eind', 'Totaalbedrag', 'Betaald']
        for col_num, header in enumerate(headers, 1):
            cell = ws_facturen.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Data
        for row_num, factuur in enumerate(facturen.data, 2):
            ws_facturen.cell(row=row_num, column=1).value = factuur.get('id')
            ws_facturen.cell(row=row_num, column=2).value = factuur.get('factuur_nummer')
            ws_facturen.cell(row=row_num, column=3).value = factuur.get('datum')
            ws_facturen.cell(row=row_num, column=4).value = factuur.get('clienten', {}).get('naam')
            ws_facturen.cell(row=row_num, column=5).value = factuur.get('start_datum')
            ws_facturen.cell(row=row_num, column=6).value = factuur.get('eind_datum')
            ws_facturen.cell(row=row_num, column=7).value = factuur.get('totaal_bedrag')
            ws_facturen.cell(row=row_num, column=8).value = 'Ja' if factuur.get('betaald') else 'Nee'
        
        # Maak overzicht werkblad
        ws_overzicht = wb.create_sheet("Overzicht Facturen", 0)  # Plaats vooraan
        
        # Titel
        periode_titel = f"Overzicht facturen {start_datum} t/m {eind_datum}"
        ws_overzicht.merge_cells('A1:H1')
        titel_cell = ws_overzicht.cell(row=1, column=1)
        titel_cell.value = periode_titel
        titel_cell.font = Font(size=14, bold=True)
        titel_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Factuurnummer', 'Datum', 'Client', 'Totaalbedrag', 'Betaald']
        for col_num, header in enumerate(headers, 1):
            cell = ws_overzicht.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Data
        for row_num, factuur in enumerate(facturen.data, 4):
            ws_overzicht.cell(row=row_num, column=1).value = factuur.get('factuur_nummer')
            ws_overzicht.cell(row=row_num, column=2).value = factuur.get('datum')
            ws_overzicht.cell(row=row_num, column=3).value = factuur.get('clienten', {}).get('naam')
            ws_overzicht.cell(row=row_num, column=4).value = factuur.get('totaal_bedrag')
            ws_overzicht.cell(row=row_num, column=5).value = 'Ja' if factuur.get('betaald') else 'Nee'
        
        # Totale rij
        total_row = len(facturen.data) + 4
        ws_overzicht.cell(row=total_row, column=3).value = "TOTAAL:"
        ws_overzicht.cell(row=total_row, column=3).font = Font(bold=True)
        
        totaal_formule = f"=SUM(D4:D{total_row-1})"
        ws_overzicht.cell(row=total_row, column=4).value = totaal_formule
        ws_overzicht.cell(row=total_row, column=4).font = Font(bold=True)
        
        # Pas kolombreedtes aan
        for ws in [ws_overzicht, ws_clienten, ws_behandeling_types, ws_behandelingen, ws_facturen]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sla Excel op in geheugen en stuur als response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Maak bestandsnaam met periode
        periode_str = f"{start_datum.replace('-', '')}_tm_{eind_datum.replace('-', '')}"
        filename = f"Facturen_{periode_str}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API Routes voor factuursjablonen
@app.route('/api/factuur_sjablonen', methods=['GET'])
def get_factuur_sjablonen():
    try:
        response = supabase.table('factuur_sjablonen').select('*').execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/factuur_sjablonen', methods=['POST'])
def add_factuur_sjabloon():
    try:
        sjabloon_data = request.json
        
        # Als deze sjabloon actief is, zet alle andere op inactief
        if sjabloon_data.get('actief'):
            supabase.table('factuur_sjablonen').update({'actief': False}).neq('id', -1).execute()
        
        response = supabase.table('factuur_sjablonen').insert(sjabloon_data).execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/factuur_sjablonen/<id>', methods=['PUT'])
def update_factuur_sjabloon(id):
    try:
        sjabloon_data = request.json
        
        # Als deze sjabloon actief is, zet alle andere op inactief
        if sjabloon_data.get('actief'):
            supabase.table('factuur_sjablonen').update({'actief': False}).neq('id', id).execute()
        
        response = supabase.table('factuur_sjablonen').update(sjabloon_data).eq('id', id).execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/factuur_sjablonen/<id>', methods=['DELETE'])
def delete_factuur_sjabloon(id):
    try:
        response = supabase.table('factuur_sjablonen').delete().eq('id', id).execute()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/factuur_sjablonen/<id>/activeer', methods=['PUT'])
def activeer_factuur_sjabloon(id):
    try:
        # Zet alle sjablonen op inactief
        supabase.table('factuur_sjablonen').update({'actief': False}).neq('id', -1).execute()
        
        # Zet de gekozen sjabloon op actief
        response = supabase.table('factuur_sjablonen').update({'actief': True}).eq('id', id).execute()
        return jsonify(response.data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)